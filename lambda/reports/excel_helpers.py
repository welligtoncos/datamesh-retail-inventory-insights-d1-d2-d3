"""openpyxl helpers — paridade visual com notebook §3."""
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

BLUE = "2563EB"
LT = "DBEAFE"


def _title(ws: Worksheet, title: str, subtitle: str, ncol: int, color: str = BLUE) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor=color)
    c2.alignment = Alignment(horizontal="center", vertical="center")


def _insight(ws: Worksheet, text: str, ncol: int, bg: str = LT) -> None:
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    c = ws.cell(row=3, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 36


def _header(ws: Worksheet, row: int, headers: list[str], color: str = BLUE) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")


def _cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    fmt: str | None = None,
    bold: bool = False,
    align: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = Font(bold=True)
    if align:
        cell.alignment = Alignment(horizontal=align)
