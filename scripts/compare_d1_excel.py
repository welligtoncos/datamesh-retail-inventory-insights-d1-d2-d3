#!/usr/bin/env python3
"""E5-US03 · Compare Excel D-1 local vs AWS (top 3 + totais)."""
from __future__ import annotations

import argparse
import sys
import tempfile
from pathlib import Path

import boto3
import pandas as pd
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "lambda" / "reports"))

from gerar_relatorio_d1 import report_s3_key  # noqa: E402


def read_data_rows(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    r = 7
    while True:
        prod = ws.cell(row=r, column=1).value
        if prod is None or prod == "TOTAL":
            break
        rows.append(
            {
                "Product ID": str(prod),
                "Category": str(ws.cell(row=r, column=2).value or ""),
                "unidades": int(ws.cell(row=r, column=3).value or 0),
                "receita": float(ws.cell(row=r, column=4).value or 0),
            }
        )
        r += 1
    return pd.DataFrame(rows)


def totals_from_agg(agg: pd.DataFrame) -> tuple[int, float]:
    return int(agg["unidades"].sum()), float(agg["receita"].sum())


def compare(local_df: pd.DataFrame, aws_df: pd.DataFrame) -> list[str]:
    errors: list[str] = []
    if len(local_df) != len(aws_df):
        errors.append(f"row count: local={len(local_df)} aws={len(aws_df)}")

    lu, lr = totals_from_agg(local_df)
    au, ar = totals_from_agg(aws_df)
    if lu != au:
        errors.append(f"total unidades: local={lu} aws={au}")
    if abs(lr - ar) > 0.01:
        errors.append(f"total receita: local={lr:.2f} aws={ar:.2f}")

    for i in range(min(3, len(local_df), len(aws_df))):
        lp = local_df.iloc[i]
        ap = aws_df.iloc[i]
        if lp["Product ID"] != ap["Product ID"]:
            errors.append(
                f"top{i+1} product: local={lp['Product ID']} aws={ap['Product ID']}"
            )
        if int(lp["unidades"]) != int(ap["unidades"]):
            errors.append(
                f"top{i+1} unidades: local={lp['unidades']} aws={ap['unidades']}"
            )
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Paridade Excel D-1 local vs S3")
    parser.add_argument("--data-execucao", default="2022-01-02")
    parser.add_argument("--dia-dado", default="2022-01-01")
    parser.add_argument("--bucket", default="retail-inventory-insights-dev-use1")
    parser.add_argument("--region", default="us-east-1")
    parser.add_argument("--local", default=None, help="Local xlsx path")
    args = parser.parse_args()

    local_path = Path(
        args.local
        or REPO / f"relatorio_D1_exec{args.data_execucao}_dado{args.dia_dado}.xlsx"
    )
    if not local_path.exists():
        print(f"FAIL: local xlsx missing: {local_path}", file=sys.stderr)
        return 1

    key = report_s3_key(args.data_execucao, args.dia_dado)
    tmp = Path(tempfile.gettempdir()) / Path(key).name
    boto3.client("s3", region_name=args.region).download_file(args.bucket, key, str(tmp))

    local_df = read_data_rows(local_path)
    aws_df = read_data_rows(tmp)

    # Ordenar como agg (por unidades desc) para comparar ranking
    local_df = local_df.sort_values("unidades", ascending=False).reset_index(drop=True)
    aws_df = aws_df.sort_values("unidades", ascending=False).reset_index(drop=True)

    errors = compare(local_df, aws_df)
    if errors:
        print("PARITY FAIL:")
        for e in errors:
            print(f"  - {e}")
        return 1

    print("PARITY OK")
    lu, lr = totals_from_agg(local_df)
    print(f"  produtos={len(local_df)} unidades={lu:,} receita={lr:,.2f}")
    print(f"  top3: {', '.join(local_df.head(3)['Product ID'].tolist())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
